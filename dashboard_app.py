import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import logging
import base64
from pathlib import Path
import tempfile
import os
import io
import time

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Define expected columns and their types
EXPECTED_COLUMNS = {
    'MUNICIPIO': 'string',
    'BAIRRO': 'string',
    'TIPO_CADASTRO': 'string',
    'SITUACAO_LIGACAO': 'string',
    'IRREGULARIDADE_IDENTIFICADA': 'string',
    'SITUACAO_HIDROMETRO': 'string',
    'TIPO_EDIFICACAO': 'string',
    'FONTE_ALTERNATIVA': 'string',
    'TOTAL_DE_MORADORES': 'float64',
    'QUANTOS_LITROS_TOTAIS': 'float64',
    'PADRAO_DO_IMOVEL': 'string',
    'LOGRADOURO': 'string',
    'QUADRA': 'string',
    'TIPO_VISITA': 'string'
}

# Função para carregar e converter logo em base64
@st.cache_data
def get_logo_base64():
    try:
        logo_path = Path("logo-aguas-do-para.png")
        if logo_path.exists():
            with open(logo_path, "rb") as f:
                data = f.read()
                return base64.b64encode(data).decode()
    except Exception as e:
        logger.error(f"Erro ao carregar logo: {str(e)}")
    return None

def _save_uploaded_file(uploaded_file) -> str:
    """Salva uploaded_file (Streamlit UploadedFile) em arquivo temporário e retorna o caminho."""
    try:
        suffix = Path(uploaded_file.name).suffix
        tmp_dir = tempfile.gettempdir()
        tmp_src = os.path.join(tmp_dir, f"uploaded_{int(time.time())}{suffix}")
        with open(tmp_src, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return tmp_src
    except Exception as e:
        logger.error(f"Erro ao salvar arquivo enviado: {e}")
        raise


# Processa o DataFrame carregado (conversões e otimizações)
def process_loaded_df(df: pd.DataFrame) -> pd.DataFrame:
    try:
        # Processar colunas
        for col, dtype in EXPECTED_COLUMNS.items():
            if col in df.columns:
                try:
                    if dtype == 'string':
                        df[col] = df[col].fillna('Não informado').astype('string')
                    elif dtype == 'float64':
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                except Exception as e:
                    logger.error(f"Erro ao converter coluna {col}: {str(e)}")

        # Criar coluna auxiliar para hidrômetro
        if 'SITUACAO_HIDROMETRO' in df.columns:
            df['POSSUI_HIDROMETRO'] = df['SITUACAO_HIDROMETRO'].apply(
                lambda x: 'SIM' if pd.notna(x) and str(x).upper() in ['NORMAL', 'QUEBRADO', 'INSTALADO'] else 'NÃO'
            )

        # Garantir que colunas numéricas existam
        if 'TOTAL_DE_MORADORES' not in df.columns:
            df['TOTAL_DE_MORADORES'] = 0
        if 'QUANTOS_LITROS_TOTAIS' not in df.columns:
            df['QUANTOS_LITROS_TOTAIS'] = 0

        # --- Otimizações de tipos para reduzir uso de memória e payload ---
        try:
            category_cols = [
                'MUNICIPIO','BAIRRO','STATUS','SITUACAO_LIGACAO',
                'TIPO_VISITA','PADRAO_DO_IMOVEL','IRREGULARIDADE_IDENTIFICADA','TIPO_EDIFICACAO'
            ]
            for col in category_cols:
                if col in df.columns:
                    df[col] = df[col].astype('category')

            # downcast em colunas numéricas conhecidas
            for col in ['TOTAL_DE_MORADORES','QUANTOS_LITROS_TOTAIS']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('float32')
        except Exception as e:
            logger.debug(f"Não foi possível otimizar tipos: {e}")
            pass

        return df
    except Exception as e:
        logger.error(f"Erro ao processar DataFrame: {e}")
        return df


# Load the data from a given path (cached by path)
@st.cache_data
def load_data(path: str) -> pd.DataFrame:
    try:
        if path.lower().endswith('.parquet'):
            df = pd.read_parquet(path)
        elif path.lower().endswith('.csv'):
            df = pd.read_csv(path)
        else:
            # assume excel
            df = pd.read_excel(path, engine='openpyxl')

        logger.info(f"Dados carregados de {path}: {len(df)} registros")
        return process_loaded_df(df)
    except Exception as e:
        logger.error(f"Erro ao carregar dados de {path}: {e}")
        st.error(f"Erro ao carregar dados: {e}")
        return pd.DataFrame()
    try:
        # Carregar diretamente do Excel
        df = pd.read_excel(
            "BASE_GERAL_2025_10_25 VS01.xlsx",
            engine='openpyxl'
        )
        
        logger.info(f"Dados carregados: {len(df)} registros")
        
        # Processar colunas
        for col, dtype in EXPECTED_COLUMNS.items():
            if col in df.columns:
                try:
                    if dtype == 'string':
                        df[col] = df[col].fillna('Não informado').astype('string')
                    elif dtype == 'float64':
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                except Exception as e:
                    logger.error(f"Erro ao converter coluna {col}: {str(e)}")

        # Criar coluna auxiliar para hidrômetro
        if 'SITUACAO_HIDROMETRO' in df.columns:
            df['POSSUI_HIDROMETRO'] = df['SITUACAO_HIDROMETRO'].apply(
                lambda x: 'SIM' if pd.notna(x) and str(x).upper() in ['NORMAL', 'QUEBRADO', 'INSTALADO'] else 'NÃO'
            )
        
        # Garantir que colunas numéricas existam
        if 'TOTAL_DE_MORADORES' not in df.columns:
            df['TOTAL_DE_MORADORES'] = 0
        if 'QUANTOS_LITROS_TOTAIS' not in df.columns:
            df['QUANTOS_LITROS_TOTAIS'] = 0
        # --- Otimizações de tipos para reduzir uso de memória e payload ---
        try:
            category_cols = [
                'MUNICIPIO','BAIRRO','STATUS','SITUACAO_LIGACAO',
                'TIPO_VISITA','PADRAO_DO_IMOVEL','IRREGULARIDADE_IDENTIFICADA','TIPO_EDIFICACAO'
            ]
            for col in category_cols:
                if col in df.columns:
                    # converter para category reduz consideravelmente o tamanho em memória
                    df[col] = df[col].astype('category')

            # downcast em colunas numéricas conhecidas
            for col in ['TOTAL_DE_MORADORES','QUANTOS_LITROS_TOTAIS']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('float32')
        except Exception as e:
            logger.debug(f"Não foi possível otimizar tipos: {e}")
            pass
            
        return df

    except Exception as e:
        logger.error(f"Erro ao carregar dados: {str(e)}")
        st.error(f"Erro ao carregar dados: {str(e)}")
        return pd.DataFrame()

# logo
logo_base64 = get_logo_base64()

# --- Input de arquivo pelo usuário (sidebar) ---
st.sidebar.markdown("""
    <div style="background: linear-gradient(135deg, #E3F2FD 0%, #BBDEFB 100%); 
                padding: 0.8rem; 
                border-radius: 8px; 
                margin-bottom: 1rem;
                border: 1px solid #90CAF9;">
        <p style="font-size: 0.9rem; font-weight: 600; color: #1976D2; margin: 0 0 0.5rem 0;">
            📁 Upload de Dados
        </p>
    </div>
""", unsafe_allow_html=True)

uploaded_file = st.sidebar.file_uploader(
    "Arquivo",
    type=['xlsx','xls','csv'],
    label_visibility="collapsed"
)

if uploaded_file is not None:
    try:
        # Mostrar loading enquanto processa
        with st.spinner(''):
            tmp_src = _save_uploaded_file(uploaded_file)
            st.session_state['uploaded_path'] = tmp_src
            st.session_state['file_uploaded'] = True
        st.sidebar.success("✓ Carregado", icon="✅")
    except Exception as e:
        st.sidebar.error(f"Erro: {e}")

# Determinar caminho dos dados
data_path = st.session_state.get('uploaded_path', None)
if data_path:
    # Mostrar loading apenas na primeira vez que carrega
    if st.session_state.get('file_uploaded', False):
        with st.spinner('Processando dados...'):
            df = load_data(data_path)
        st.session_state['file_uploaded'] = False
    else:
        df = load_data(data_path)
else:
    st.sidebar.info("💡 Faça upload da planilha", icon="ℹ️")
    df = pd.DataFrame()

st.sidebar.markdown("<br>", unsafe_allow_html=True)


def prepare_map_points(df_map, max_points=5000):
    """Reduz o número de pontos para o mapa por amostragem ou agregação espacial.
    Retorna (df_plot, info) onde info é uma string com mensagem quando ocorreu redução.
    """
    try:
        n = len(df_map)
        if n <= max_points:
            return df_map.copy(), None

        # Tentar agregar por grid (arredondando coordenadas) para reduzir pontos
        df_tmp = df_map.copy()
        df_tmp['lat_r'] = df_tmp['LATITUDE'].round(3)
        df_tmp['lon_r'] = df_tmp['LONGITUDE'].round(3)
        df_agg = df_tmp.groupby(['lat_r', 'lon_r', 'STATUS']).size().reset_index(name='count')
        df_agg = df_agg.rename(columns={'lat_r': 'LATITUDE', 'lon_r': 'LONGITUDE'})

        info = f"Dados do mapa reduzidos: {n} -> {len(df_agg)} pontos (agregação por grid)."
        return df_agg, info
    except Exception as e:
        logger.debug(f"prepare_map_points falhou: {e}")
        # fallback: amostragem simples
        try:
            df_samp = df_map.sample(n=min(len(df_map), max_points), random_state=42)
            info = f"Dados do mapa amostrados para {len(df_samp)} pontos (fallback)."
            return df_samp, info
        except Exception:
            return df_map.copy(), None

# Configuração da página
st.set_page_config(
    page_title="Painel Diário de Cadastros",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Função para mostrar loading animado
def show_loading():
    st.markdown("""
        <style>
        .loading-overlay {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: linear-gradient(135deg, rgba(25, 118, 210, 0.95) 0%, rgba(33, 150, 243, 0.95) 100%);
            display: flex;
            justify-content: center;
            align-items: center;
            z-index: 9999;
            animation: fadeIn 0.3s ease-in;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
        
        .loading-container {
            text-align: center;
            color: white;
        }
        
        .loading-spinner {
            width: 80px;
            height: 80px;
            margin: 0 auto 30px;
            border: 6px solid rgba(255, 255, 255, 0.3);
            border-top: 6px solid white;
            border-radius: 50%;
            animation: spin 1s linear infinite;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .loading-title {
            font-size: 2rem;
            font-weight: 700;
            margin-bottom: 15px;
            animation: pulse 1.5s ease-in-out infinite;
        }
        
        .loading-subtitle {
            font-size: 1.1rem;
            font-weight: 400;
            opacity: 0.9;
            margin-bottom: 30px;
        }
        
        .loading-dots {
            display: inline-block;
        }
        
        .loading-dots span {
            animation: blink 1.4s infinite;
            animation-fill-mode: both;
        }
        
        .loading-dots span:nth-child(2) {
            animation-delay: 0.2s;
        }
        
        .loading-dots span:nth-child(3) {
            animation-delay: 0.4s;
        }
        
        @keyframes blink {
            0%, 80%, 100% { opacity: 0; }
            40% { opacity: 1; }
        }
        
        @keyframes pulse {
            0%, 100% { transform: scale(1); }
            50% { transform: scale(1.05); }
        }
        
        .loading-progress {
            width: 300px;
            height: 4px;
            background: rgba(255, 255, 255, 0.3);
            border-radius: 2px;
            margin: 20px auto;
            overflow: hidden;
        }
        
        .loading-progress-bar {
            height: 100%;
            background: white;
            border-radius: 2px;
            animation: progress 2s ease-in-out infinite;
        }
        
        @keyframes progress {
            0% { width: 0%; transform: translateX(0); }
            50% { width: 70%; transform: translateX(0); }
            100% { width: 100%; transform: translateX(0); }
        }
        </style>
        
        <div class="loading-overlay" id="loadingOverlay">
            <div class="loading-container">
                <div class="loading-spinner"></div>
                <div class="loading-title">💧 Águas do Pará</div>
                <div class="loading-subtitle">
                    Carregando dados
                    <span class="loading-dots">
                        <span>.</span><span>.</span><span>.</span>
                    </span>
                </div>
                <div class="loading-progress">
                    <div class="loading-progress-bar"></div>
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)

def hide_loading():
    st.markdown("""
        <script>
        const overlay = document.getElementById('loadingOverlay');
        if (overlay) {
            overlay.style.animation = 'fadeOut 0.3s ease-out';
            setTimeout(() => {
                overlay.style.display = 'none';
            }, 300);
        }
        </script>
        <style>
        @keyframes fadeOut {
            from { opacity: 1; }
            to { opacity: 0; }
        }
        </style>
    """, unsafe_allow_html=True)

# CSS moderno e clean - Tema Claro
st.markdown("""
    <style>
    /* Importar fonte moderna */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
    }
    
    :root {
        --primary-color: #1976D2;
        --secondary-color: #FF9800;
        --accent-color: #66BB6A;
        --background: #F5F7FA;
        --surface: #FFFFFF;
        --surface-light: #F8F9FA;
        --text: #1a1a1a;
        --text-light: #64748b;
        --border: #E2E8F0;
        --shadow: rgba(0, 0, 0, 0.08);
    }
    
    /* Background principal */
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #F5F7FA 0%, #E8EEF5 100%);
    }
    
    /* Sidebar moderna */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #FFFFFF 0%, #F8F9FA 100%);
        border-right: 1px solid var(--border);
        box-shadow: 2px 0 12px var(--shadow);
    }
    
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h2 {
        color: var(--primary-color) !important;
        font-weight: 600;
        font-size: 1.1rem;
        margin-bottom: 1.5rem;
    }
    
    /* Estilizar labels e textos da sidebar */
    [data-testid="stSidebar"] label {
        color: var(--text) !important;
        font-weight: 500;
    }
    
    [data-testid="stSidebar"] .stMarkdown {
        color: var(--text-light);
    }
    
    /* Header com logo */
    .dashboard-header {
        background: linear-gradient(135deg, #FFFFFF 0%, #F0F4F8 100%);
        border-radius: 16px;
        padding: 2rem;
        margin-bottom: 2rem;
        box-shadow: 0 4px 20px var(--shadow);
        border: 1px solid var(--border);
        display: flex;
        align-items: center;
        gap: 2rem;
    }
    
    .logo-container {
        flex-shrink: 0;
    }
    
    .logo-container img {
        height: 80px;
        width: auto;
        object-fit: contain;
    }
    
    .header-content {
        flex: 1;
    }
    
    .header-title {
        color: var(--primary-color);
        font-size: 2rem;
        font-weight: 700;
        margin: 0;
        line-height: 1.2;
    }
    
    .header-subtitle {
        color: var(--text-light);
        font-size: 1rem;
        font-weight: 400;
        margin-top: 0.5rem;
    }
    
    /* Cards de métricas modernos */
    .metric-card {
        background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFB 100%);
        border-radius: 16px;
        padding: 1.5rem;
        border: 1px solid var(--border);
        box-shadow: 0 2px 12px var(--shadow);
        transition: all 0.3s ease;
        height: 100%;
    }
    
    .metric-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 8px 24px rgba(25, 118, 210, 0.15);
        border-color: var(--primary-color);
    }
    
    .metric-icon {
        width: 48px;
        height: 48px;
        background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
        border-radius: 12px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 24px;
        margin-bottom: 1rem;
        box-shadow: 0 4px 12px rgba(79, 195, 247, 0.3);
    }
    
    .metric-label {
        color: var(--text-light);
        font-size: 0.875rem;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 0.5rem;
    }
    
    .metric-value {
        color: var(--text);
        font-size: 2rem;
        font-weight: 700;
        line-height: 1;
    }
    
    /* Títulos de seção */
    .section-title {
        color: var(--primary-color);
        font-size: 1.5rem;
        font-weight: 600;
        margin: 2rem 0 1rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid var(--primary-color);
    }
    
    /* Gráficos */
    .js-plotly-plot {
        border-radius: 12px;
        overflow: hidden;
        background: white;
        box-shadow: 0 2px 12px var(--shadow);
        border: 1px solid var(--border);
    }
    
    /* Tabs modernos */
    [data-testid="stTabs"] {
        background: transparent;
    }
    
    [data-testid="stTabs"] [role="tablist"] {
        background: white;
        border-radius: 12px;
        padding: 0.5rem;
        border: 1px solid var(--border);
        box-shadow: 0 2px 8px var(--shadow);
    }
    
    [data-testid="stTabs"] [role="tab"] {
        color: var(--text-light);
        background: transparent;
        border: none;
        border-radius: 8px;
        padding: 0.75rem 1.5rem;
        font-weight: 500;
        transition: all 0.2s ease;
    }
    
    [data-testid="stTabs"] [role="tab"]:hover {
        background: var(--surface-light);
        color: var(--text);
    }
    
    [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
        background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
        color: white !important;
    }
    
    /* Divisores */
    hr {
        border: none;
        height: 1px;
        background: linear-gradient(90deg, transparent, var(--border), transparent);
        margin: 2rem 0;
    }
    
    /* Footer */
    .footer-stats {
        background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFB 100%);
        border-radius: 12px;
        padding: 1.5rem;
        margin-top: 2rem;
        border: 1px solid var(--border);
        box-shadow: 0 2px 12px var(--shadow);
        display: flex;
        justify-content: space-around;
        align-items: center;
    }
    
    .footer-stat {
        text-align: center;
    }
    
    .footer-stat-label {
        color: var(--text-light);
        font-size: 0.875rem;
        font-weight: 500;
        margin-bottom: 0.25rem;
    }
    
    .footer-stat-value {
        color: var(--primary-color);
        font-size: 1.25rem;
        font-weight: 700;
    }
    
    /* Multiselect customizado */
    [data-baseweb="select"] {
        border-radius: 8px;
    }
    
    /* Remover padding extra */
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    
    /* Responsividade */
    @media (max-width: 768px) {
        .dashboard-header {
            flex-direction: column;
            text-align: center;
        }
        
        .header-title {
            font-size: 1.5rem;
        }
        
        .logo-container img {
            height: 60px;
        }
    }
    </style>
""", unsafe_allow_html=True)

# Header com logo
if logo_base64:
    st.markdown(f"""
        <div class="dashboard-header">
            <div class="logo-container">
                <img src="data:image/png;base64,{logo_base64}" alt="Logo Águas do Pará">
            </div>
            <div class="header-content">
                <h1 class="header-title">Painel Diário de Cadastros</h1>
                <p class="header-subtitle">Análise completa dos dados de abastecimento de água municipal</p>
            </div>
        </div>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
        <div class="dashboard-header">
            <div class="header-content">
                <h1 class="header-title">💧 Painel Diário de Cadastros</h1>
                <p class="header-subtitle">Análise completa dos dados de abastecimento de água municipal</p>
            </div>
        </div>
    """, unsafe_allow_html=True)

# --- Sidebar Filters ---
st.sidebar.markdown('<p style="font-size: 1.2rem; font-weight: 600; color: #1976D2; margin-bottom: 1.5rem;">🔍 Filtros</p>', unsafe_allow_html=True)

if not df.empty:
    # Município
    st.sidebar.markdown('<p style="font-size: 0.85rem; font-weight: 500; color: #64748b; margin-bottom: 0.3rem;">Município</p>', unsafe_allow_html=True)
    municipio_options = sorted(df["MUNICIPIO"].unique())
    
    col_m1, col_m2 = st.sidebar.columns(2)
    with col_m1:
        if st.button("✓ Todos", key="todos_mun", use_container_width=True):
            st.session_state.municipios = municipio_options
    with col_m2:
        if st.button("✗ Limpar", key="limpar_mun", use_container_width=True):
            st.session_state.municipios = []
    
    if 'municipios' not in st.session_state:
        st.session_state.municipios = municipio_options
    
    selected_municipio = st.sidebar.multiselect(
        "Selecione",
        options=municipio_options,
        default=st.session_state.municipios,
        key="select_mun",
        label_visibility="collapsed"
    )
    st.session_state.municipios = selected_municipio
    
    st.sidebar.markdown("---")
    
    # Bairro
    st.sidebar.markdown('<p style="font-size: 0.85rem; font-weight: 500; color: #64748b; margin-bottom: 0.3rem;">Bairro</p>', unsafe_allow_html=True)
    bairro_options = sorted(df[df["MUNICIPIO"].isin(selected_municipio)]["BAIRRO"].unique())
    
    col_b1, col_b2 = st.sidebar.columns(2)
    with col_b1:
        if st.button("✓ Todos", key="todos_bairro", use_container_width=True):
            st.session_state.bairros = bairro_options
    with col_b2:
        if st.button("✗ Limpar", key="limpar_bairro", use_container_width=True):
            st.session_state.bairros = []
    
    if 'bairros' not in st.session_state:
        st.session_state.bairros = bairro_options
    
    selected_bairro = st.sidebar.multiselect(
        "Selecione",
        options=bairro_options,
        default=st.session_state.bairros,
        key="select_bairro",
        label_visibility="collapsed"
    )
    st.session_state.bairros = selected_bairro
    
    st.sidebar.markdown("---")

    # Filtro por STATUS
    if "STATUS" in df.columns:
        st.sidebar.markdown('<p style="font-size: 0.85rem; font-weight: 500; color: #64748b; margin-bottom: 0.3rem;">Status</p>', unsafe_allow_html=True)
        status_options = sorted([s for s in df["STATUS"].unique() if pd.notna(s)])
        
        col_s1, col_s2 = st.sidebar.columns(2)
        with col_s1:
            if st.button("✓ Todos", key="todos_status", use_container_width=True):
                st.session_state.status = status_options
        with col_s2:
            if st.button("✗ Limpar", key="limpar_status", use_container_width=True):
                st.session_state.status = []
        
        if 'status' not in st.session_state:
            st.session_state.status = status_options
        
        selected_status = st.sidebar.multiselect(
            "Selecione",
            options=status_options,
            default=st.session_state.status,
            key="select_status",
            label_visibility="collapsed"
        )
        st.session_state.status = selected_status
        
        df_selection = df[
            df["MUNICIPIO"].isin(selected_municipio) & 
            df["BAIRRO"].isin(selected_bairro) &
            df["STATUS"].isin(selected_status)
        ]
    else:
        df_selection = df[df["MUNICIPIO"].isin(selected_municipio) & df["BAIRRO"].isin(selected_bairro)]
else:
    st.error("Não foi possível carregar os dados.")
    st.stop()

# --- KPI Metrics Section ---
st.markdown('<div class="section-title">📈 Métricas Principais</div>', unsafe_allow_html=True)
col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)

with col_kpi1:
    total_clientes = len(df_selection)
    total_clientes_fmt = f"{total_clientes:,.0f}".replace(',', '.')
    st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">👥</div>
            <div class="metric-label">Total de Clientes</div>
            <div class="metric-value">{total_clientes_fmt}</div>
        </div>
    """, unsafe_allow_html=True)

with col_kpi2:
    if 'POSSUI_HIDROMETRO' in df_selection.columns:
        perc_hidrometro = (df_selection['POSSUI_HIDROMETRO'].value_counts(normalize=True).get('SIM', 0) * 100)
    else:
        perc_hidrometro = 0
    st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">💧</div>
            <div class="metric-label">Com Hidrômetro</div>
            <div class="metric-value">{perc_hidrometro:.1f}%</div>
        </div>
    """, unsafe_allow_html=True)

with col_kpi3:
    # Calcular média apenas de imóveis com 1 ou mais moradores
    if 'TOTAL_DE_MORADORES' in df_selection.columns:
        df_com_moradores = df_selection[df_selection['TOTAL_DE_MORADORES'] >= 1]
        media_moradores = df_com_moradores['TOTAL_DE_MORADORES'].mean() if len(df_com_moradores) > 0 else 0
    else:
        media_moradores = 0
    media_moradores_fmt = f"{media_moradores:.1f}".replace('.', ',')
    st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">🏠</div>
            <div class="metric-label">Média de Moradores</div>
            <div class="metric-value">{media_moradores_fmt}</div>
        </div>
    """, unsafe_allow_html=True)

with col_kpi4:
    # Contar matrículas duplicadas (excluindo valores 0)
    if 'MATRICULA' in df_selection.columns:
        # Filtrar matrículas diferentes de zero
        df_matriculas_validas = df_selection[df_selection['MATRICULA'] != 0].copy()
        
        # Contar todas as ocorrências de cada matrícula
        contagem_matriculas = df_matriculas_validas['MATRICULA'].value_counts()
        
        # Selecionar apenas as matrículas que aparecem mais de uma vez
        matriculas_duplicadas = contagem_matriculas[contagem_matriculas > 1].sum() - len(contagem_matriculas[contagem_matriculas > 1])
        
        # Criar DataFrame com todas as matrículas duplicadas e seus dados completos
        matriculas_dup_list = contagem_matriculas[contagem_matriculas > 1].index.tolist()
        df_duplicadas_completo = df_matriculas_validas[df_matriculas_validas['MATRICULA'].isin(matriculas_dup_list)].copy()
        
        # Ordenar por matrícula para facilitar visualização
        df_duplicadas_completo = df_duplicadas_completo.sort_values('MATRICULA')
        
        # Adicionar coluna com contagem de ocorrências
        df_duplicadas_completo['TOTAL_OCORRENCIAS'] = df_duplicadas_completo['MATRICULA'].map(contagem_matriculas)
        
        # Debug info
        st.sidebar.markdown("### Informações de Validação")
        with st.sidebar.expander("Detalhes de Matrículas Duplicadas"):
            st.write("Matrículas que aparecem mais de uma vez:")
            duplicatas = contagem_matriculas[contagem_matriculas > 1].to_dict()
            for mat, count in duplicatas.items():
                st.write(f"Matrícula {mat}: {count} ocorrências")
            
            st.markdown("---")
            
            # Botão para baixar relatório completo de duplicidades
            if not df_duplicadas_completo.empty:
                st.markdown("#### 📥 Baixar Relatório de Duplicidades")
                
                # Preparar DataFrame organizado para exportação
                df_export = df_duplicadas_completo.copy()
                
                # Reorganizar colunas para melhor visualização
                colunas_prioritarias = ['MATRICULA', 'TOTAL_OCORRENCIAS', 'MUNICIPIO', 'BAIRRO', 'LOGRADOURO', 
                                       'TIPO_CADASTRO', 'STATUS', 'SITUACAO_LIGACAO', 'REAMBULADOR']
                
                # Adicionar colunas prioritárias que existem no DataFrame
                colunas_ordenadas = [col for col in colunas_prioritarias if col in df_export.columns]
                
                # Adicionar demais colunas
                outras_colunas = [col for col in df_export.columns if col not in colunas_ordenadas]
                colunas_finais = colunas_ordenadas + outras_colunas
                
                df_export = df_export[colunas_finais]
                
                # Criar arquivo Excel em memória
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='Matrículas Duplicadas')
                    
                    # Obter a planilha para formatação
                    workbook = writer.book
                    worksheet = writer.sheets['Matrículas Duplicadas']
                    
                    # Formatar cabeçalho
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF', size=11)
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Aplicar estilo ao cabeçalho
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                    
                    # Ajustar largura das colunas
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Aplicar bordas e alinhamento aos dados
                    thin_border = Border(
                        left=Side(style='thin', color='E2E8F0'),
                        right=Side(style='thin', color='E2E8F0'),
                        top=Side(style='thin', color='E2E8F0'),
                        bottom=Side(style='thin', color='E2E8F0')
                    )
                    
                    data_alignment = Alignment(horizontal='left', vertical='center')
                    
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                                   min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = data_alignment
                    
                    # Destacar coluna TOTAL_OCORRENCIAS
                    if 'TOTAL_OCORRENCIAS' in df_export.columns:
                        col_idx = df_export.columns.get_loc('TOTAL_OCORRENCIAS') + 1
                        highlight_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
                        
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.fill = highlight_fill
                            cell.font = Font(bold=True, color='E65100')
                    
                    # Congelar primeira linha
                    worksheet.freeze_panes = 'A2'
                
                excel_data = output.getvalue()
                
                st.download_button(
                    label="📊 Baixar Relatório Excel Formatado",
                    data=excel_data,
                    file_name=f'relatorio_matriculas_duplicadas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    help=f"Baixar relatório formatado com {len(df_duplicadas_completo)} registros de {len(matriculas_dup_list)} matrículas duplicadas",
                    use_container_width=True
                )
                
                st.info(f"📊 Total: {len(df_duplicadas_completo)} registros de {len(matriculas_dup_list)} matrículas duplicadas")
    else:
        matriculas_duplicadas = 0
    matriculas_duplicadas_fmt = f"{matriculas_duplicadas:,.0f}".replace(',', '.')
    st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">🚨</div>
            <div class="metric-label">Matrículas Duplicadas</div>
            <div class="metric-value">{matriculas_duplicadas_fmt}</div>
        </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# --- Gráfico de STATUS e Mapa ---
st.markdown('<div class="section-title">📋 Distribuição por Status</div>', unsafe_allow_html=True)

col_status = st.container()

with col_status:
    if not df_selection.empty and "STATUS" in df_selection.columns:
        status_dist = df_selection["STATUS"].value_counts().reset_index()
        status_dist.columns = ["Status", "Contagem"]
        
        # Definir cores vibrantes específicas para cada status
        color_map = {
            'COLETAR': '#4FC3F7',  # Azul vibrante
            'VALIDAR': '#FFB74D',  # Laranja vibrante
            'REAMBULADO': '#66BB6A',  # Verde vibrante
            'AGUARDANDO CADASTRO': '#FF9800',  # Laranja escuro
            'EXCLUIR': '#78909C',  # Cinza azulado
            'CORRIGIR': '#EF5350'  # Vermelho vibrante
        }
        
        # Criar lista de cores na ordem dos status
        colors = [color_map.get(status.upper(), '#808080') for status in status_dist['Status']]
        
        # Adicionar informação de valor absoluto ao texto
        status_dist['Texto'] = status_dist.apply(
            lambda x: f"{x['Status']}<br>{x['Contagem']:,} registros<br>{(x['Contagem']/status_dist['Contagem'].sum()*100):.1f}%", 
            axis=1
        )
        
        fig_status = px.pie(
            status_dist,
            values="Contagem",
            names="Status",
            title="Distribuição por Status do Cadastro",
            color_discrete_sequence=colors,
            hole=0.4,
            custom_data=['Texto']  # incluir texto customizado
        )
        fig_status.update_layout(
            template="plotly_white",
            paper_bgcolor="rgba(255, 255, 255, 0.95)",
            plot_bgcolor="rgba(255, 255, 255, 0.95)",
            font=dict(color="#1a1a1a", size=14, family="Inter"),
            title_font=dict(size=18, color="#1976D2", family="Inter"),
            margin=dict(t=60, b=40, l=40, r=40),
            height=500,
            showlegend=True,
            legend=dict(
                font=dict(color="#64748b"),
                bgcolor="rgba(255, 255, 255, 0.8)"
            )
        )
        fig_status.update_traces(
            textposition='inside',
            texttemplate="%{value:,.0f}<br>(%{percent:.1%})",
            textinfo='text',
            hovertemplate="%{customdata[0]}<extra></extra>",
            textfont_size=12
        )
        st.plotly_chart(fig_status, use_container_width=True, key="status_chart")

# Mapa removido por opção do usuário (pode pesar muito).

st.markdown("---")

# --- Análise Técnica Completa ---
st.markdown('<div class="section-title">📊 Análise Técnica e Operacional</div>', unsafe_allow_html=True)

# Row 1 - Gráficos de Pizza (3 colunas)
col1, col2, col3 = st.columns(3)

with col1:
    if not df_selection.empty and "SITUACAO_LIGACAO" in df_selection.columns:
        ligacao_dist = df_selection["SITUACAO_LIGACAO"].value_counts().reset_index()
        ligacao_dist.columns = ["Status", "Contagem"]
        # Adicionar informação de valor absoluto ao texto
        ligacao_dist['Texto'] = ligacao_dist.apply(
            lambda x: f"{x['Status']}<br>{x['Contagem']:,} registros<br>{(x['Contagem']/ligacao_dist['Contagem'].sum()*100):.1f}%", 
            axis=1
        )
        
        fig_ligacao = px.pie(
            ligacao_dist,
            values="Contagem",
            names="Status",
            title="Status de Ligação",
            color_discrete_sequence=["#4FC3F7", "#FF9800", "#66BB6A", "#FFB74D", "#AB47BC", "#EF5350"],
            hole=0.4,
            custom_data=['Texto']
        )
        fig_ligacao.update_layout(
            template="plotly_white",
            paper_bgcolor="rgba(255, 255, 255, 0.95)",
            plot_bgcolor="rgba(255, 255, 255, 0.95)",
            font=dict(color="#1a1a1a", size=12, family="Inter"),
            title_font=dict(size=16, color="#1976D2", family="Inter"),
            margin=dict(t=50, b=20, l=20, r=20),
            showlegend=True,
            legend=dict(font=dict(color="#64748b"))
        )
        fig_ligacao.update_traces(
            textposition='inside',
            texttemplate="%{value:,.0f}<br>(%{percent:.1%})",
            textinfo='text',
            hovertemplate="%{customdata[0]}<extra></extra>",
            textfont_size=12
        )
        st.plotly_chart(fig_ligacao, use_container_width=True, key="ligacao_chart")

with col2:
    if not df_selection.empty and "IRREGULARIDADE_IDENTIFICADA" in df_selection.columns:
        irregularidade_dist = df_selection["IRREGULARIDADE_IDENTIFICADA"].value_counts().reset_index()
        irregularidade_dist.columns = ["Irregularidade", "Contagem"]
        # Adicionar informação de valor absoluto ao texto
        irregularidade_dist['Texto'] = irregularidade_dist.apply(
            lambda x: f"{x['Irregularidade']}<br>{x['Contagem']:,} registros<br>{(x['Contagem']/irregularidade_dist['Contagem'].sum()*100):.1f}%", 
            axis=1
        )
        
        fig_irregularidade = px.pie(
            irregularidade_dist,
            values="Contagem",
            names="Irregularidade",
            title="Irregularidades Identificadas",
            color_discrete_sequence=["#EF5350", "#66BB6A", "#FFB74D", "#4FC3F7", "#AB47BC"],
            hole=0.4,
            custom_data=['Texto']
        )
        fig_irregularidade.update_layout(
            template="plotly_white",
            paper_bgcolor="rgba(255, 255, 255, 0.95)",
            plot_bgcolor="rgba(255, 255, 255, 0.95)",
            font=dict(color="#1a1a1a", size=12, family="Inter"),
            title_font=dict(size=16, color="#1976D2", family="Inter"),
            margin=dict(t=50, b=20, l=20, r=20),
            showlegend=True,
            legend=dict(font=dict(color="#64748b"))
        )
        fig_irregularidade.update_traces(
            textposition='inside',
            texttemplate="%{value:,.0f}<br>(%{percent:.1%})",
            textinfo='text',
            hovertemplate="%{customdata[0]}<extra></extra>",
            textfont_size=12
        )
        st.plotly_chart(fig_irregularidade, use_container_width=True, key="irregularidade_chart")

with col3:
    if not df_selection.empty and "TIPO_EDIFICACAO" in df_selection.columns:
        edificacao_dist = df_selection["TIPO_EDIFICACAO"].value_counts().reset_index()
        edificacao_dist.columns = ["Tipo", "Contagem"]
        # Adicionar informação de valor absoluto ao texto
        edificacao_dist['Texto'] = edificacao_dist.apply(
            lambda x: f"{x['Tipo']}<br>{x['Contagem']:,} registros<br>{(x['Contagem']/edificacao_dist['Contagem'].sum()*100):.1f}%", 
            axis=1
        )
        
        fig_edificacao = px.pie(
            edificacao_dist,
            values="Contagem",
            names="Tipo",
            title="Tipo de Edificação",
            color_discrete_sequence=["#4FC3F7", "#FFB74D", "#66BB6A", "#FF9800", "#AB47BC", "#EF5350"],
            hole=0.4,
            custom_data=['Texto']
        )
        fig_edificacao.update_layout(
            template="plotly_white",
            paper_bgcolor="rgba(255, 255, 255, 0.95)",
            plot_bgcolor="rgba(255, 255, 255, 0.95)",
            font=dict(color="#1a1a1a", size=12, family="Inter"),
            title_font=dict(size=16, color="#1976D2", family="Inter"),
            margin=dict(t=50, b=20, l=20, r=20),
            showlegend=True,
            legend=dict(font=dict(color="#64748b"))
        )
        fig_edificacao.update_traces(
            textposition='inside',
            texttemplate="%{value:,.0f}<br>(%{percent:.1%})",
            textinfo='text',
            hovertemplate="%{customdata[0]}<extra></extra>",
            textfont_size=12
        )
        st.plotly_chart(fig_edificacao, use_container_width=True, key="edificacao_chart")

st.markdown("---")

# Row 2 - Padrão de Edificação (1 coluna)
if not df_selection.empty and "PADRAO_EDIFICACAO" in df_selection.columns:
    padrao_edif_dist = df_selection["PADRAO_EDIFICACAO"].value_counts().reset_index()
    padrao_edif_dist.columns = ["Padrão", "Contagem"]
    
    # Criar gráfico de radar (rede)
    fig_padrao_edif = go.Figure()
    
    fig_padrao_edif.add_trace(go.Scatterpolar(
        r=padrao_edif_dist["Contagem"],
        theta=padrao_edif_dist["Padrão"],
        fill='toself',
        fillcolor='rgba(79, 195, 247, 0.3)',
        line=dict(color='#4FC3F7', width=2),
        marker=dict(size=8, color='#1976D2'),
        name='Padrão de Edificação',
        hovertemplate="<b>%{theta}</b><br>Quantidade: %{r}<extra></extra>"
    ))
    
    fig_padrao_edif.update_layout(
        template="plotly_white",
        paper_bgcolor="rgba(255, 255, 255, 0.95)",
        plot_bgcolor="rgba(255, 255, 255, 0.95)",
        font=dict(color="#1a1a1a", size=12, family="Inter"),
        title="Padrão de Edificação",
        title_font=dict(size=16, color="#1976D2", family="Inter"),
        polar=dict(
            radialaxis=dict(
                visible=True,
                gridcolor="#E2E8F0",
                linecolor="#E2E8F0"
            ),
            angularaxis=dict(
                gridcolor="#E2E8F0",
                linecolor="#E2E8F0"
            ),
            bgcolor="rgba(255, 255, 255, 0.95)"
        ),
        showlegend=False,
        margin=dict(t=60, b=40, l=60, r=60),
        height=400
    )
    st.plotly_chart(fig_padrao_edif, use_container_width=True, key="padrao_edif_chart")

st.markdown("---")

# Row 3 - Economias e Quadra (2 colunas)
col5, col6_quadra = st.columns(2)

with col5:
    # Gráfico de Economias (RES, COM, PUB, IND) - Colunas Múltiplas
    economias_cols = ['ECONOMIA_RES', 'ECONOMIA_COM', 'ECONOMIA_PUB', 'ECONOMIA_IND']
    
    # Verificar se as colunas existem
    cols_existentes = [col for col in economias_cols if col in df_selection.columns]
    
    if cols_existentes:
        fig_economias = go.Figure()
        
        # Cores para cada tipo
        cores = {
            'ECONOMIA_RES': '#4FC3F7',
            'ECONOMIA_COM': '#FF9800',
            'ECONOMIA_PUB': '#66BB6A',
            'ECONOMIA_IND': '#AB47BC'
        }
        
        nomes = {
            'ECONOMIA_RES': 'Residencial',
            'ECONOMIA_COM': 'Comercial',
            'ECONOMIA_PUB': 'Público',
            'ECONOMIA_IND': 'Industrial'
        }
        
        # Adicionar uma barra para cada tipo de economia
        for col in cols_existentes:
            total = df_selection[col].sum()
            tipo = nomes.get(col, col.replace('ECONOMIA_', ''))
            
            fig_economias.add_trace(go.Bar(
                name=tipo,
                x=[tipo],
                y=[total],
                marker_color=cores.get(col, '#808080'),
                text=[f"{total:,.0f}".replace(',', '.')],
                textposition='outside',
                hovertemplate=f"<b>{tipo}</b><br>Total: {total:,.0f}".replace(',', '.') + "<extra></extra>"
            ))
        
        fig_economias.update_layout(
            template="plotly_white",
            paper_bgcolor="rgba(255, 255, 255, 0.95)",
            plot_bgcolor="rgba(255, 255, 255, 0.95)",
            font=dict(color="#1a1a1a", size=12, family="Inter"),
            title="Economias por Tipo",
            title_font=dict(size=16, color="#1976D2", family="Inter"),
            xaxis_title="Tipo de Economia",
            yaxis_title="Quantidade",
            showlegend=True,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1,
                font=dict(color="#64748b")
            ),
            xaxis=dict(gridcolor="#E2E8F0"),
            yaxis=dict(gridcolor="#E2E8F0"),
            margin=dict(t=100, b=40, l=40, r=20),
            height=450,
            barmode='group'
        )
        
        st.plotly_chart(fig_economias, use_container_width=True, key="economias_chart")

with col6_quadra:
    if not df_selection.empty and "QUADRA" in df_selection.columns:
        # Filtrar para remover "Não informado"
        df_quadra_filtrado = df_selection[
            (df_selection["QUADRA"].notna()) & 
            (df_selection["QUADRA"] != "Não informado") &
            (df_selection["QUADRA"].str.strip() != "")
        ]
        
        if not df_quadra_filtrado.empty:
            quadra_dist = df_quadra_filtrado["QUADRA"].value_counts().reset_index()
            quadra_dist.columns = ["Quadra", "Quantidade"]
            
            # Converter quadra para string sem decimais (remover .0)
            quadra_dist["Quadra"] = quadra_dist["Quadra"].apply(
                lambda x: str(int(float(x))) if str(x).replace('.', '').replace('-', '').isdigit() else str(x)
            )
            
            quadra_dist = quadra_dist.sort_values("Quadra")
            
            fig_quadra = px.bar(
                quadra_dist,
                x="Quadra",
                y="Quantidade",
                title="Quantitativo de Clientes por Quadra",
                color_discrete_sequence=["#FF9800"]
            )
            fig_quadra.update_layout(
                template="plotly_white",
                paper_bgcolor="rgba(255, 255, 255, 0.95)",
                plot_bgcolor="rgba(255, 255, 255, 0.95)",
                font=dict(color="#1a1a1a", size=12, family="Inter"),
                title_font=dict(size=16, color="#1976D2", family="Inter"),
                xaxis_title="Quadra",
                yaxis_title="Quantidade",
                showlegend=False,
                xaxis=dict(type="category", gridcolor="#E2E8F0"),
                yaxis=dict(gridcolor="#E2E8F0"),
                margin=dict(t=50, b=40, l=40, r=20)
            )
            st.plotly_chart(fig_quadra, use_container_width=True, key="quadra_chart")

st.markdown("---")

# Row 4 - Tipo de Visita e Padrão do Imóvel (2 colunas)
col7, col8 = st.columns(2)

with col7:
    if not df_selection.empty and "TIPO_VISITA" in df_selection.columns:
        # Filtrar para remover "Não informado"
        df_visita_filtrado = df_selection[
            (df_selection["TIPO_VISITA"].notna()) & 
            (df_selection["TIPO_VISITA"] != "Não informado") &
            (df_selection["TIPO_VISITA"].str.strip() != "")
        ]
        
        if not df_visita_filtrado.empty:
            visita_dist = df_visita_filtrado["TIPO_VISITA"].value_counts().reset_index()
            visita_dist.columns = ["Tipo de Visita", "Quantidade"]
            
            fig_visita = go.Figure(data=[
                go.Bar(
                    y=visita_dist["Tipo de Visita"],
                    x=visita_dist["Quantidade"],
                    orientation="h",
                    marker=dict(
                        color="#66BB6A",
                        line=dict(color="#4FC3F7", width=2)
                    ),
                    text=visita_dist["Quantidade"],
                    textposition="outside",
                    textfont=dict(color="#FFFFFF"),
                    hovertemplate="<b>%{y}</b><br>Quantidade: %{x}<extra></extra>"
                )
            ])
            fig_visita.update_layout(
                template="plotly_white",
                paper_bgcolor="rgba(255, 255, 255, 0.95)",
                plot_bgcolor="rgba(255, 255, 255, 0.95)",
                font=dict(color="#1a1a1a", size=12, family="Inter"),
                title="Distribuição por Tipo de Visita",
                title_font=dict(size=16, color="#1976D2", family="Inter"),
                xaxis_title="Quantidade",
                yaxis_title="Tipo",
                height=400,
                showlegend=False,
                yaxis={"categoryorder": "total ascending", "gridcolor": "#E2E8F0"},
                xaxis=dict(gridcolor="#E2E8F0"),
                margin=dict(t=50, b=40, l=20, r=40)
            )
            st.plotly_chart(fig_visita, use_container_width=True, key="visita_chart")

with col8:
    if not df_selection.empty and "PADRAO_DO_IMOVEL" in df_selection.columns:
        # Filtrar para remover "Não informado"
        df_padrao_filtrado = df_selection[
            (df_selection["PADRAO_DO_IMOVEL"].notna()) & 
            (df_selection["PADRAO_DO_IMOVEL"] != "Não informado") &
            (df_selection["PADRAO_DO_IMOVEL"].str.strip() != "")
        ]
        
        if not df_padrao_filtrado.empty:
            padrao_dist = df_padrao_filtrado["PADRAO_DO_IMOVEL"].value_counts().reset_index()
            padrao_dist.columns = ["Padrão", "Quantidade"]
            
            # Criar gráfico de radar (rede) para Padrão do Imóvel
            fig_padrao = go.Figure()
            
            fig_padrao.add_trace(go.Scatterpolar(
                r=padrao_dist["Quantidade"],
                theta=padrao_dist["Padrão"],
                fill='toself',
                fillcolor='rgba(255, 183, 77, 0.3)',
                line=dict(color='#FFB74D', width=2),
                marker=dict(size=8, color='#FF9800'),
                name='Padrão do Imóvel',
                hovertemplate="<b>%{theta}</b><br>Quantidade: %{r}<extra></extra>"
            ))
            
            fig_padrao.update_layout(
                template="plotly_white",
                paper_bgcolor="rgba(255, 255, 255, 0.95)",
                plot_bgcolor="rgba(255, 255, 255, 0.95)",
                font=dict(color="#1a1a1a", size=12, family="Inter"),
                title="Distribuição por Padrão do Imóvel",
                title_font=dict(size=16, color="#1976D2", family="Inter"),
                polar=dict(
                    radialaxis=dict(
                        visible=True,
                        gridcolor="#E2E8F0",
                        linecolor="#E2E8F0"
                    ),
                    angularaxis=dict(
                        gridcolor="#E2E8F0",
                        linecolor="#E2E8F0"
                    ),
                    bgcolor="rgba(255, 255, 255, 0.95)"
                ),
                showlegend=False,
                margin=dict(t=60, b=40, l=60, r=60),
                height=400
            )
            st.plotly_chart(fig_padrao, use_container_width=True, key="padrao_chart")

st.markdown("---")

# --- Nova Aba de Produtividade ---
st.markdown('<div class="section-title">🏆 Produtividade de Reambuladores</div>', unsafe_allow_html=True)

# Filtros de data
col_filtro1, col_filtro2, col_filtro3 = st.columns([2, 2, 2])

with col_filtro1:
    tipo_filtro = st.selectbox(
        "Tipo de Filtro",
        ["Hoje", "Por Período", "Por Mês"],
        key="tipo_filtro_reambulador"
    )

# Converter DATA_COLETA para datetime se existir
if 'DATA_COLETA' in df_selection.columns:
    df_selection_copy = df_selection.copy()
    df_selection_copy['DATA_COLETA'] = pd.to_datetime(df_selection_copy['DATA_COLETA'], errors='coerce')
    hoje = date.today()
    hoje = pd.Timestamp.now().date()
    
    if tipo_filtro == "Hoje":
        # Filtrar para registros de hoje
        df_filtrado = df_selection_copy[
            df_selection_copy['DATA_COLETA'].dt.date == hoje
        ]
        with col_filtro2:
            st.info(f"Mostrando dados de hoje ({hoje.strftime('%d/%m/%Y')})")
            
    elif tipo_filtro == "Por Período":
        with col_filtro2:
            # Obter datas mínima e máxima
            data_min = df_selection_copy['DATA_COLETA'].min()
            data_max = df_selection_copy['DATA_COLETA'].max()
            
            if pd.notna(data_min) and pd.notna(data_max):
                data_inicio = st.date_input(
                    "Data Início",
                    value=data_min.date(),
                    min_value=data_min.date(),
                    max_value=data_max.date(),
                    key="data_inicio_reambulador"
                )
        
        with col_filtro3:
            if pd.notna(data_min) and pd.notna(data_max):
                data_fim = st.date_input(
                    "Data Fim",
                    value=data_max.date(),
                    min_value=data_min.date(),
                    max_value=data_max.date(),
                    key="data_fim_reambulador"
                )
                
                # Filtrar por período
                df_filtrado = df_selection_copy[
                    (df_selection_copy['DATA_COLETA'].dt.date >= data_inicio) &
                    (df_selection_copy['DATA_COLETA'].dt.date <= data_fim)
                ]
            else:
                st.warning("Não há dados de data disponíveis")
                df_filtrado = df_selection_copy
    else:  # Por Mês
        with col_filtro2:
            # Criar lista de meses disponíveis
            df_selection_copy['ANO_MES'] = df_selection_copy['DATA_COLETA'].dt.to_period('M')
            meses_disponiveis = sorted(df_selection_copy['ANO_MES'].dropna().unique(), reverse=True)
            
            if len(meses_disponiveis) > 0:
                mes_selecionado = st.selectbox(
                    "Selecione o Mês",
                    options=meses_disponiveis,
                    format_func=lambda x: x.strftime('%B/%Y') if pd.notna(x) else 'N/A',
                    key="mes_reambulador"
                )
                
                # Filtrar por mês
                df_filtrado = df_selection_copy[df_selection_copy['ANO_MES'] == mes_selecionado]
            else:
                st.warning("Não há dados de data disponíveis")
                df_filtrado = df_selection_copy
else:
    st.warning("Coluna DATA_COLETA não encontrada nos dados")
    df_filtrado = df_selection

st.markdown("---")

# Criar ranking de reambuladores
if 'REAMBULADOR' in df_filtrado.columns and not df_filtrado.empty:
    # Contar quantas vezes cada reambulador aparece (número de visitas)
    ranking = df_filtrado['REAMBULADOR'].value_counts().reset_index()
    ranking.columns = ['Reambulador', 'Número de Visitas']
    
    # Converter para string e remover valores nulos ou vazios
    ranking['Reambulador'] = ranking['Reambulador'].astype(str)
    ranking = ranking[ranking['Reambulador'].notna() & (ranking['Reambulador'] != '') & (ranking['Reambulador'] != 'nan')]
    
    # Remover valores nulos ou vazios
    ranking = ranking[ranking['Reambulador'].notna() & (ranking['Reambulador'] != '')]
    
    if not ranking.empty:
        # Métricas principais
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        
        # Calcular média de visitas
        media_visitas = ranking['Número de Visitas'].mean()
        
        with col_m1:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">👥</div>
                    <div class="metric-label">Total de Reambuladores</div>
                    <div class="metric-value">{len(ranking)}</div>
                </div>
            """, unsafe_allow_html=True)
        
        with col_m2:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">📊</div>
                    <div class="metric-label">Total de Visitas</div>
                    <div class="metric-value">{ranking['Número de Visitas'].sum():,}</div>
                </div>
            """, unsafe_allow_html=True)
        
        with col_m3:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">📈</div>
                    <div class="metric-label">Média de Visitas</div>
                    <div class="metric-value">{media_visitas:.1f}</div>
                </div>
            """, unsafe_allow_html=True)
            
        with col_m4:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">🏆</div>
                    <div class="metric-label">Melhor Performance</div>
                    <div class="metric-value">{ranking.iloc[0]['Número de Visitas']:,} visitas</div>
                </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Gráfico de ranking e tabela
        col_grafico, col_tabela = st.columns([2, 1])
        
        with col_grafico:
            # Gráfico de barras horizontal
            fig_ranking = go.Figure(data=[
                go.Bar(
                    y=ranking['Reambulador'],
                    x=ranking['Número de Visitas'],
                    orientation='h',
                    marker=dict(
                        color=ranking['Número de Visitas'],
                        colorscale=[[0, '#E6F7FF'], [0.5, '#A8DADC'], [1, '#7BC8D4']],
                        line=dict(color='#A8DADC', width=1)
                    ),
                    text=ranking['Número de Visitas'],
                    textposition='outside',
                    hovertemplate="<b>%{y}</b><br>Visitas: %{x}<extra></extra>"
                )
            ])
            
            fig_ranking.update_layout(
                template="plotly_white",
                paper_bgcolor="rgba(255, 255, 255, 0.95)",
                plot_bgcolor="rgba(255, 255, 255, 0.95)",
                font=dict(color="#1a1a1a", size=12, family="Inter"),
                title="Ranking de Reambuladores por Número de Visitas",
                title_font=dict(size=18, color="#1976D2", family="Inter"),
                xaxis_title="Número de Visitas",
                yaxis_title="Reambulador",
                height=max(400, len(ranking) * 25),
                showlegend=False,
                yaxis={"categoryorder": "total ascending", "gridcolor": "#E2E8F0"},
                xaxis=dict(gridcolor="#E2E8F0"),
                margin=dict(t=60, b=40, l=200, r=40)
            )
            
            st.plotly_chart(fig_ranking, use_container_width=True, key="ranking_reambulador")
        
        with col_tabela:
            st.markdown("#### 📋 Ranking Completo")
            
            # Adicionar informações de horário da primeira e última coleta
            if 'DATA_COLETA' in df_filtrado.columns:
                horarios_info = []
                for reambulador in ranking['Reambulador']:
                    df_reambulador = df_filtrado[df_filtrado['REAMBULADOR'].astype(str) == reambulador]
                    
                    # Obter primeira e última coleta
                    datas_validas = df_reambulador['DATA_COLETA'].dropna()
                    
                    if len(datas_validas) > 0:
                        primeira_coleta = datas_validas.min()
                        ultima_coleta = datas_validas.max()
                        
                        # Extrair apenas a hora
                        hora_primeira = primeira_coleta.strftime('%H:%M') if pd.notna(primeira_coleta) else 'N/A'
                        hora_ultima = ultima_coleta.strftime('%H:%M') if pd.notna(ultima_coleta) else 'N/A'
                        
                        horarios_info.append(f"{hora_primeira} - {hora_ultima}")
                    else:
                        horarios_info.append('N/A')
                
                # Criar coluna com nome e horários
                ranking_display = ranking.copy()
                ranking_display['Horários'] = horarios_info
            else:
                ranking_display = ranking.copy()
            
            # Adicionar posição no ranking
            ranking_display.insert(0, 'Posição', range(1, len(ranking_display) + 1))
            
            # Mostrar apenas top 100 na tabela para evitar envio massivo ao frontend
            max_show = 100
            
            # Aplicar estilo para destacar horários em vermelho
            def highlight_horarios(row):
                if 'Horários' in row.index:
                    return [''] * (len(row) - 1) + ['color: #EF5350; font-weight: bold;']
                return [''] * len(row)
            
            if 'Horários' in ranking_display.columns:
                st.dataframe(
                    ranking_display.head(max_show).style.apply(highlight_horarios, axis=1),
                    hide_index=True,
                    use_container_width=True,
                    height=max(400, min(len(ranking), max_show) * 35 + 38)
                )
            else:
                st.dataframe(
                    ranking_display.head(max_show),
                    hide_index=True,
                    use_container_width=True,
                    height=max(400, min(len(ranking), max_show) * 35 + 38)
                )

            # Oferecer download do CSV completo
            try:
                csv_bytes = ranking.to_csv(index=False).encode('utf-8')
                st.download_button("📥 Baixar CSV (completo)", csv_bytes, file_name='ranking_reambulador.csv', mime='text/csv')
            except Exception:
                # se algo falhar, apenas permitir copiar por clipboard
                st.markdown("Não foi possível gerar o arquivo para download.")
    else:
        st.warning("Não há dados de reambuladores para o período selecionado")
else:
    st.warning("Coluna REAMBULADOR não encontrada nos dados ou não há dados disponíveis")

# Footer com estatísticas
total_registros_fmt = f"{len(df_selection):,.0f}".replace(',', '.')
st.markdown(f"""
    <div class="footer-stats">
        <div class="footer-stat">
            <div class="footer-stat-label">Total de Registros</div>
            <div class="footer-stat-value">{total_registros_fmt}</div>
        </div>
        <div class="footer-stat">
            <div class="footer-stat-label">Última Atualização</div>
            <div class="footer-stat-value">{datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
        </div>
    </div>
""", unsafe_allow_html=True)
